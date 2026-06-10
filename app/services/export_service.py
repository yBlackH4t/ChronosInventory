"""
Serviço de exportação (Excel) para API.
Responsabilidade: gerar arquivo de produtos para download.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Tuple, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.stock_service import StockService
from core.constants import DATE_FORMAT_FILE
from core.utils.file_utils import FileUtils
from core.database.connection import DatabaseConnection
from core.database.repositories.inventory_location_repository import InventoryLocationRepository
from app.models.inventory_location import InventoryLocation

class ExportService:
    def __init__(self) -> None:
        self.stock_service = StockService()

    def _get_active_locations(self) -> List[InventoryLocation]:
        conn = DatabaseConnection().get_connection()
        try:
            repo = InventoryLocationRepository(conn)
            return repo.get_all_active()
        finally:
            conn.close()

    def export_products_excel(self) -> Tuple[str, int]:
        products = self.stock_service.get_all_products()
        active_locations = self._get_active_locations()
        
        temp_dir = FileUtils.get_temp_directory()
        timestamp = datetime.now().strftime(DATE_FORMAT_FILE)
        filename = f"export_produtos_{timestamp}.xlsx"
        path = os.path.join(temp_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"

        headers = ["ID", "Produto"]
        for loc in active_locations:
            headers.append(f"Estoque {loc.label}")
        headers.extend(["Total", "Onde tem"])

        ws.append(headers)

        for product in products:
            row = [product.id, product.nome]
            total = 0
            has_stock_locs = []
            for loc in active_locations:
                qty = product.inventories.get(loc.id, 0)
                row.append(qty)
                total += qty
                if qty > 0:
                    has_stock_locs.append(loc.label)
            row.append(total)
            
            where = " / ".join(has_stock_locs) if has_stock_locs else "Sem saldo"
            row.append(where)
            
            ws.append(row)

        wb.save(path)
        return path, len(products)

    def export_stock_overview_excel(self) -> Tuple[str, int]:
        products = self.stock_service.get_all_products()
        active_locations = self._get_active_locations()
        
        temp_dir = FileUtils.get_temp_directory()
        timestamp = datetime.now().strftime(DATE_FORMAT_FILE)
        filename = f"estoque_resumo_{timestamp}.xlsx"
        path = os.path.join(temp_dir, filename)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        ws_stock = wb.create_sheet("Estoque")

        total_items = len(products)
        items_with_stock = 0
        total_global = 0
        
        loc_totals = {loc.id: 0 for loc in active_locations}
        
        for product in products:
            prod_total = 0
            for loc in active_locations:
                qty = product.inventories.get(loc.id, 0)
                loc_totals[loc.id] += qty
                prod_total += qty
            total_global += prod_total
            if prod_total > 0:
                items_with_stock += 1

        title_fill = PatternFill("solid", fgColor="1D4ED8")
        subtitle_fill = PatternFill("solid", fgColor="DBEAFE")
        header_fill = PatternFill("solid", fgColor="E0E7FF")
        accent_fill = PatternFill("solid", fgColor="EEF2FF")
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"] = "Chronos Inventory - Resumo de Estoque"
        ws_summary["A1"].font = Font(color="FFFFFF", bold=True, size=15)
        ws_summary["A1"].fill = title_fill
        ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.merge_cells("A2:D2")
        ws_summary["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_summary["A2"].font = Font(color="1E293B", italic=True, size=11)
        ws_summary["A2"].fill = subtitle_fill
        ws_summary["A2"].alignment = Alignment(horizontal="center", vertical="center")

        summary_rows = [
            ("Total de produtos ativos", total_items),
            ("Itens com estoque", items_with_stock),
        ]
        
        for loc in active_locations:
            summary_rows.append((f"Total de pecas em {loc.label}", loc_totals[loc.id]))
            
        summary_rows.append(("Total global de pecas", total_global))

        start_row = 4
        for index, (label, value) in enumerate(summary_rows, start=start_row):
            ws_summary[f"A{index}"] = label
            ws_summary[f"A{index}"].font = Font(bold=True, color="334155")
            ws_summary[f"A{index}"].fill = accent_fill
            ws_summary[f"A{index}"].border = border
            ws_summary[f"B{index}"] = value
            ws_summary[f"B{index}"].font = Font(bold=True, color="0F172A")
            ws_summary[f"B{index}"].border = border

        info_row_start = start_row + len(summary_rows) + 2
        
        ws_summary[f"A{info_row_start}"] = "Leitura rapida"
        ws_summary[f"A{info_row_start}"].font = Font(bold=True, color="1E3A8A")
        ws_summary[f"A{info_row_start}"].fill = header_fill
        ws_summary[f"A{info_row_start}"].border = border
        ws_summary[f"A{info_row_start+1}"] = "Use a aba Estoque para ver item por item, com os estoques de cada local, total e onde existe saldo."
        ws_summary[f"A{info_row_start+1}"].alignment = Alignment(wrap_text=True)
        ws_summary[f"A{info_row_start+1}"].border = border
        ws_summary.merge_cells(f"A{info_row_start+1}:D{info_row_start+1}")

        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 18
        ws_summary.column_dimensions["C"].width = 18
        ws_summary.column_dimensions["D"].width = 18

        stock_headers = ["ID", "Produto"]
        for loc in active_locations:
            stock_headers.append(f"Estoque {loc.label}")
        stock_headers.extend(["Total", "Onde tem"])
        
        for col_index, header in enumerate(stock_headers, start=1):
            cell = ws_stock.cell(row=1, column=col_index)
            cell.value = header
            cell.font = Font(bold=True, color="1E293B")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, product in enumerate(products, start=2):
            total = 0
            has_stock_locs = []
            loc_values = []
            
            for loc in active_locations:
                qty = product.inventories.get(loc.id, 0)
                loc_values.append(qty)
                total += qty
                if qty > 0:
                    has_stock_locs.append(loc.label)
                    
            where = " / ".join(has_stock_locs) if has_stock_locs else "Sem saldo"
            
            values = [product.id, product.nome] + loc_values + [total, where]
            
            for col_index, value in enumerate(values, start=1):
                cell = ws_stock.cell(row=row_index, column=col_index)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if col_index > 2 and col_index < len(values):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_index % 2 == 0:
                for col_index in range(1, len(stock_headers) + 1):
                    ws_stock.cell(row=row_index, column=col_index).fill = PatternFill("solid", fgColor="F8FAFC")

        ws_stock.freeze_panes = "A2"
        ws_stock.auto_filter.ref = f"A1:{get_column_letter(len(stock_headers))}{max(len(products) + 1, 2)}"

        ws_stock.column_dimensions["A"].width = 10
        ws_stock.column_dimensions["B"].width = 42
        
        for i in range(len(active_locations)):
            ws_stock.column_dimensions[get_column_letter(3 + i)].width = 16
            
        ws_stock.column_dimensions[get_column_letter(3 + len(active_locations))].width = 10
        ws_stock.column_dimensions[get_column_letter(4 + len(active_locations))].width = 18

        wb.save(path)
        return path, total_items

