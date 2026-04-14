from datetime import datetime
import io
import json
import os
import sqlite3
import socket
import tempfile
import zipfile

import pandas as pd
from openpyxl import load_workbook


def _create_product(client, nome="Produto Teste", qtd_canoas=1, qtd_pf=0):
    payload = {"nome": nome, "qtd_canoas": qtd_canoas, "qtd_pf": qtd_pf}
    response = client.post("/produtos", json=payload)
    assert response.status_code == 201
    return response.json()["data"]["id"]


def _find_free_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return sock.getsockname()[1]


def test_health_success(client):
    response = client.get("/health")
    assert response.status_code == 200
    body = response.json()
    assert "data" in body
    assert "meta" in body
    assert body["data"]["status"] == "ok"
    assert response.headers.get("X-Request-ID")


def test_list_products_success(client):
    response = client.get("/produtos")
    assert response.status_code == 200
    body = response.json()
    assert "data" in body
    assert "meta" in body
    assert isinstance(body["data"], list)
    assert set(body["meta"].keys()) == {
        "page",
        "page_size",
        "total_items",
        "total_pages",
        "has_next",
    }


def test_list_products_can_sort_by_total_stock(client):
    low_id = _create_product(client, "Produto Estoque Baixo", qtd_canoas=1, qtd_pf=0)
    high_id = _create_product(client, "Produto Estoque Alto", qtd_canoas=3, qtd_pf=4)

    desc = client.get("/produtos?sort=-total_stock&page=1&page_size=20")
    assert desc.status_code == 200
    desc_ids = [item["id"] for item in desc.json()["data"]]
    assert desc_ids.index(high_id) < desc_ids.index(low_id)

    asc = client.get("/produtos?sort=total_stock&page=1&page_size=20")
    assert asc.status_code == 200
    asc_ids = [item["id"] for item in asc.json()["data"]]
    assert asc_ids.index(low_id) < asc_ids.index(high_id)


def test_list_products_can_search_by_hash_id_and_status_screen(client):
    target_id = _create_product(client, "4031196", qtd_canoas=1, qtd_pf=0)
    other_id = _create_product(client, "5610131", qtd_canoas=2, qtd_pf=0)

    by_name = client.get("/produtos?query=4031196&page=1&page_size=20")
    assert by_name.status_code == 200
    by_name_ids = [item["id"] for item in by_name.json()["data"]]
    assert target_id in by_name_ids

    by_id = client.get(f"/produtos?query=%23{target_id}&page=1&page_size=20")
    assert by_id.status_code == 200
    by_id_data = by_id.json()["data"]
    assert [item["id"] for item in by_id_data] == [target_id]

    status_by_id = client.get(f"/produtos/gestao-status?query=%23{other_id}&page=1&page_size=20")
    assert status_by_id.status_code == 200
    status_data = status_by_id.json()["data"]
    assert [item["id"] for item in status_data] == [other_id]


def test_stock_profiles_lifecycle(client):
    listed = client.get("/sistema/estoques")
    assert listed.status_code == 200
    state = listed.json()["data"]
    assert state["active_profile_id"] == "default"
    assert state["restart_required"] is False
    assert isinstance(state["profiles"], list)
    assert any(item["id"] == "default" for item in state["profiles"])

    created = client.post(
        "/sistema/estoques",
        json={"name": "Filial Teste", "profile_id": "filial_teste"},
    )
    assert created.status_code == 201
    created_data = created.json()["data"]
    assert created_data["id"] == "filial_teste"
    assert created_data["is_active"] is False

    switched = client.put("/sistema/estoques/ativo", json={"profile_id": "filial_teste"})
    assert switched.status_code == 200
    switched_data = switched.json()["data"]
    assert switched_data["active_profile_id"] == "filial_teste"
    assert switched_data["requires_restart"] is True

    listed_after = client.get("/sistema/estoques")
    assert listed_after.status_code == 200
    state_after = listed_after.json()["data"]
    assert state_after["active_profile_id"] == "filial_teste"
    assert state_after["restart_required"] is True
    assert any(item["id"] == "filial_teste" for item in state_after["profiles"])

    switched_back = client.put("/sistema/estoques/ativo", json={"profile_id": "default"})
    assert switched_back.status_code == 200
    switched_back_data = switched_back.json()["data"]
    assert switched_back_data["active_profile_id"] == "default"
    assert switched_back_data["requires_restart"] is True


def test_stock_profile_delete_rules(client):
    created = client.post(
        "/sistema/estoques",
        json={"name": "Filial Excluir", "profile_id": "filial_excluir"},
    )
    assert created.status_code == 201
    created_data = created.json()["data"]
    assert created_data["id"] == "filial_excluir"

    delete_default = client.delete("/sistema/estoques/default")
    assert delete_default.status_code == 400

    switched = client.put("/sistema/estoques/ativo", json={"profile_id": "filial_excluir"})
    assert switched.status_code == 200

    delete_active = client.delete("/sistema/estoques/filial_excluir")
    assert delete_active.status_code == 400

    switched_back = client.put("/sistema/estoques/ativo", json={"profile_id": "default"})
    assert switched_back.status_code == 200

    deleted = client.delete("/sistema/estoques/filial_excluir")
    assert deleted.status_code == 200
    deleted_data = deleted.json()["data"]
    assert deleted_data["deleted_profile_id"] == "filial_excluir"
    assert not os.path.exists(created_data["path"])

    listed_after = client.get("/sistema/estoques")
    assert listed_after.status_code == 200
    ids_after = {item["id"] for item in listed_after.json()["data"]["profiles"]}
    assert "filial_excluir" not in ids_after


def test_compare_stock_databases(client):
    with tempfile.TemporaryDirectory() as tmpdir:
        left_path = os.path.join(tmpdir, "left.db")
        right_path = os.path.join(tmpdir, "right.db")

        def write_db(path: str, rows: list[tuple[int, str, int, int, int]]) -> None:
            conn = sqlite3.connect(path)
            try:
                conn.execute(
                    """
                    CREATE TABLE produtos (
                        id INTEGER PRIMARY KEY,
                        nome TEXT NOT NULL,
                        qtd_canoas INTEGER DEFAULT 0,
                        qtd_pf INTEGER DEFAULT 0,
                        ativo INTEGER NOT NULL DEFAULT 1
                    )
                    """
                )
                conn.executemany(
                    "INSERT INTO produtos (id, nome, qtd_canoas, qtd_pf, ativo) VALUES (?, ?, ?, ?, ?)",
                    rows,
                )
                conn.commit()
            finally:
                conn.close()

        write_db(
            left_path,
            [
                (1, "Produto Igual", 2, 1, 1),
                (2, "Produto Canoas", 5, 0, 1),
                (3, "Produto So Esquerda", 0, 4, 1),
            ],
        )
        write_db(
            right_path,
            [
                (1, "Produto Igual", 2, 1, 1),
                (2, "Produto Canoas", 7, 0, 1),
                (4, "Produto So Direita", 1, 0, 1),
            ],
        )

        resp = client.post(
            "/sistema/comparar-bases",
            json={
                "left_path": left_path,
                "right_path": right_path,
                "left_label": "Minha base",
                "right_label": "Base colega",
            },
        )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["left"]["label"] == "Minha base"
        assert data["right"]["label"] == "Base colega"
        assert data["summary"]["total_compared_items"] == 4
        assert data["summary"]["identical_items"] == 1
        assert data["summary"]["divergent_items"] == 3
        assert data["summary"]["only_left_items"] == 1
        assert data["summary"]["only_right_items"] == 1
        assert data["summary"]["canoas_mismatch_items"] == 1

        rows = {item["product_id"]: item for item in data["rows"]}
        assert rows[1]["statuses"] == ["IDENTICAL"]
        assert "CANOAS" in rows[2]["statuses"]
        assert "ONLY_LEFT" in rows[3]["statuses"]
        assert "ONLY_RIGHT" in rows[4]["statuses"]


def test_compare_with_published_snapshot_flow(client, tmp_path):
    share_dir = tmp_path / "compare-share"

    remote_product = _create_product(client, "Produto Snapshot Remoto", qtd_canoas=2, qtd_pf=0)
    assert remote_product > 0

    configured_remote = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "consumer",
            "official_base_dir": str(share_dir),
            "machine_label": "PC_REMOTO",
        },
    )
    assert configured_remote.status_code == 200

    published = client.post("/sistema/comparativo-publicado/publicar")
    assert published.status_code == 200
    published_data = published.json()["data"]
    assert published_data["machine_label"] == "PC_REMOTO"
    assert os.path.isfile(published_data["zip_path"])

    local_product = _create_product(client, "Produto So Local", qtd_canoas=0, qtd_pf=5)
    assert local_product > 0

    configured_local = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "consumer",
            "official_base_dir": str(share_dir),
            "machine_label": "PC_LOCAL",
        },
    )
    assert configured_local.status_code == 200

    status = client.get("/sistema/comparativo-publicado/status")
    assert status.status_code == 200
    status_data = status.json()["data"]
    assert status_data["configured"] is True
    assert any(item["machine_label"] == "PC_REMOTO" for item in status_data["available_bases"])

    compared = client.post("/sistema/comparativo-publicado/PC_REMOTO/comparar")
    assert compared.status_code == 200
    compared_data = compared.json()["data"]
    assert compared_data["right"]["label"] == "Base publicada - PC_REMOTO"
    assert compared_data["summary"]["only_left_items"] >= 1
    rows = {item["product_id"]: item for item in compared_data["rows"]}
    assert "ONLY_LEFT" in rows[local_product]["statuses"]


def test_create_and_get_product(client):
    payload = {"nome": "Produto Teste", "qtd_canoas": 1, "qtd_pf": 0}
    create = client.post("/produtos", json=payload)
    assert create.status_code == 201
    created = create.json()["data"]
    assert created["id"]

    fetched = client.get(f"/produtos/{created['id']}")
    assert fetched.status_code == 200
    fetched_body = fetched.json()["data"]
    assert fetched_body["nome"] == "PRODUTO TESTE"


def test_patch_product_observacao(client):
    product_id = _create_product(client, "Produto Obs", qtd_canoas=1, qtd_pf=0)
    resp = client.patch(f"/produtos/{product_id}", json={"observacao": "Observacao livre"})
    assert resp.status_code == 200
    body = resp.json()["data"]
    assert body["observacao"] == "Observacao livre"


def test_bulk_inactivate_hides_product_from_default_list(client):
    active_id = _create_product(client, "Produto Ativo", qtd_canoas=2, qtd_pf=0)
    inactive_id = _create_product(client, "Produto Inativo", qtd_canoas=3, qtd_pf=0)

    updated = client.put(
        "/produtos/status-lote",
        json={"ids": [inactive_id], "ativo": False, "motivo_inativacao": "Item obsoleto"},
    )
    assert updated.status_code == 200
    assert updated.json()["data"]["updated"] == 1

    default_list = client.get("/produtos")
    assert default_list.status_code == 200
    default_ids = {item["id"] for item in default_list.json()["data"]}
    assert active_id in default_ids
    assert inactive_id not in default_ids

    inactive_list = client.get("/produtos/gestao-status?status=INATIVO")
    assert inactive_list.status_code == 200
    inactive_ids = {item["id"] for item in inactive_list.json()["data"]}
    assert inactive_id in inactive_ids

    product = client.get(f"/produtos/{inactive_id}")
    assert product.status_code == 200
    assert product.json()["data"]["ativo"] is False


def test_products_status_can_filter_items_with_stock(client):
    inativo_com_estoque = _create_product(client, "Produto Inativo Com Estoque", qtd_canoas=1, qtd_pf=0)
    sem_estoque_id = _create_product(client, "Produto Sem Estoque", qtd_canoas=1, qtd_pf=0)
    _create_product(client, "Produto PF Com Estoque", qtd_canoas=0, qtd_pf=2)

    saida = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": sem_estoque_id,
            "quantidade": 1,
            "origem": "CANOAS",
        },
    )
    assert saida.status_code == 201

    updated = client.put(
        "/produtos/status-lote",
        json={"ids": [inativo_com_estoque], "ativo": False, "motivo_inativacao": "Teste filtro"},
    )
    assert updated.status_code == 200

    with_stock = client.get("/produtos/gestao-status?status=TODOS&has_stock=true")
    assert with_stock.status_code == 200
    with_stock_ids = {item["id"] for item in with_stock.json()["data"]}
    assert inativo_com_estoque in with_stock_ids
    assert any(item["qtd_pf"] > 0 for item in with_stock.json()["data"])
    assert all((item["qtd_canoas"] + item["qtd_pf"]) > 0 for item in with_stock.json()["data"])

    no_stock = client.get("/produtos/gestao-status?status=TODOS&has_stock=false")
    assert no_stock.status_code == 200
    assert all((item["qtd_canoas"] + item["qtd_pf"]) == 0 for item in no_stock.json()["data"])


def test_analytics_stock_summary_excludes_inactive_products(client):
    hidden_id = _create_product(client, "Produto Hidden", qtd_canoas=10, qtd_pf=1)
    _create_product(client, "Produto Visivel", qtd_canoas=2, qtd_pf=3)

    status = client.put("/produtos/status-lote", json={"ids": [hidden_id], "ativo": False})
    assert status.status_code == 200

    summary = client.get("/analytics/stock/summary")
    assert summary.status_code == 200
    data = summary.json()["data"]
    assert data["total_canoas"] == 2
    assert data["total_pf"] == 3
    assert data["total_geral"] == 5


def test_cannot_create_movement_for_inactive_product(client):
    product_id = _create_product(client, "Produto Bloqueado", qtd_canoas=2, qtd_pf=0)
    status = client.put("/produtos/status-lote", json={"ids": [product_id], "ativo": False})
    assert status.status_code == 200

    movement = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 1,
            "origem": "CANOAS",
        },
    )
    assert movement.status_code == 400
    body = movement.json()["error"]
    assert body["code"] == "validation_error"
    assert "Produto inativo" in body["message"]


def test_create_invalid_payload(client):
    response = client.post("/produtos", json={"nome": ""})
    assert response.status_code == 422
    body = response.json()
    assert "error" in body
    assert body["error"]["code"] == "validation_error"


def test_movement_entry_increases_stock(client):
    product_id = _create_product(client, "Produto Entrada", qtd_canoas=1, qtd_pf=0)

    movement = {
        "tipo": "ENTRADA",
        "produto_id": product_id,
        "quantidade": 3,
        "destino": "CANOAS",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 201
    assert resp.json()["data"]["produto_nome"] == "PRODUTO ENTRADA"

    product = client.get(f"/produtos/{product_id}").json()["data"]
    assert product["qtd_canoas"] == 4


def test_movement_saida_insufficient_stock(client):
    product_id = _create_product(client, "Produto Saida", qtd_canoas=1, qtd_pf=0)

    movement = {
        "tipo": "SAIDA",
        "produto_id": product_id,
        "quantidade": 2,
        "origem": "CANOAS",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 409
    body = resp.json()
    assert body["error"]["code"] == "insufficient_stock"


def test_movement_transfer_updates_both(client):
    product_id = _create_product(client, "Produto Transfer", qtd_canoas=5, qtd_pf=0)

    movement = {
        "tipo": "TRANSFERENCIA",
        "produto_id": product_id,
        "quantidade": 2,
        "origem": "CANOAS",
        "destino": "PF",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 201

    product = client.get(f"/produtos/{product_id}").json()["data"]
    assert product["qtd_canoas"] == 3
    assert product["qtd_pf"] == 2


def test_movement_saida_transferencia_externa_registers_metadata(client):
    product_id = _create_product(client, "Produto Externo", qtd_canoas=5, qtd_pf=0)

    movement = {
        "tipo": "SAIDA",
        "produto_id": product_id,
        "quantidade": 2,
        "origem": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "MATRIZ",
        "documento": "NF 123",
        "observacao": "Envio excepcional",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 201
    data = resp.json()["data"]
    assert data["natureza"] == "TRANSFERENCIA_EXTERNA"
    assert data["local_externo"] == "MATRIZ"
    assert data["documento"] == "NF 123"

    product = client.get(f"/produtos/{product_id}").json()["data"]
    assert product["qtd_canoas"] == 3


def test_movement_entrada_transferencia_externa_registers_metadata(client):
    product_id = _create_product(client, "Produto Recebido Externo", qtd_canoas=1, qtd_pf=0)

    movement = {
        "tipo": "ENTRADA",
        "produto_id": product_id,
        "quantidade": 2,
        "destino": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "MATRIZ",
        "documento": "NF REM 456",
        "observacao": "Recebido da filial",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 201
    data = resp.json()["data"]
    assert data["natureza"] == "TRANSFERENCIA_EXTERNA"
    assert data["local_externo"] == "MATRIZ"
    assert data["documento"] == "NF REM 456"

    product = client.get(f"/produtos/{product_id}").json()["data"]
    assert product["qtd_canoas"] == 3


def test_movement_devolucao_reentrada_with_reference(client):
    product_id = _create_product(client, "Produto Devolucao", qtd_canoas=1, qtd_pf=0)

    saida = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 1,
            "origem": "CANOAS",
            "documento": "NF DEV-1",
        },
    )
    assert saida.status_code == 201
    saida_id = saida.json()["data"]["id"]

    devolucao = client.post(
        "/movimentacoes",
        json={
            "tipo": "ENTRADA",
            "produto_id": product_id,
            "quantidade": 1,
            "destino": "CANOAS",
            "natureza": "DEVOLUCAO",
            "movimento_ref_id": saida_id,
            "documento": "NF DEV-1",
            "observacao": "Peca retornou",
        },
    )
    assert devolucao.status_code == 201
    devolucao_data = devolucao.json()["data"]
    assert devolucao_data["natureza"] == "DEVOLUCAO"
    assert devolucao_data["movimento_ref_id"] == saida_id

    product = client.get(f"/produtos/{product_id}").json()["data"]
    assert product["qtd_canoas"] == 1

    history = client.get(f"/produtos/{product_id}/historico").json()["data"]
    assert any(item["natureza"] == "DEVOLUCAO" for item in history)


def test_movement_invalid_devolucao_for_saida(client):
    product_id = _create_product(client, "Produto Regra", qtd_canoas=1, qtd_pf=0)
    movement = {
        "tipo": "SAIDA",
        "produto_id": product_id,
        "quantidade": 1,
        "origem": "CANOAS",
        "natureza": "DEVOLUCAO",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 400
    assert resp.json()["error"]["code"] == "validation_error"


def test_movement_ajuste_requires_reason_and_observation(client):
    product_id = _create_product(client, "Produto Ajuste", qtd_canoas=3, qtd_pf=0)
    movement = {
        "tipo": "SAIDA",
        "produto_id": product_id,
        "quantidade": 1,
        "origem": "CANOAS",
        "natureza": "AJUSTE",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 400
    assert resp.json()["error"]["code"] == "validation_error"

    ok = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 1,
            "origem": "CANOAS",
            "natureza": "AJUSTE",
            "motivo_ajuste": "PERDA",
            "observacao": "Ajuste por conferencia",
        },
    )
    assert ok.status_code == 201
    assert ok.json()["data"]["motivo_ajuste"] == "PERDA"


def test_movement_devolucao_requires_reference(client):
    product_id = _create_product(client, "Produto Sem Ref", qtd_canoas=1, qtd_pf=0)
    movement = {
        "tipo": "ENTRADA",
        "produto_id": product_id,
        "quantidade": 1,
        "destino": "CANOAS",
        "natureza": "DEVOLUCAO",
    }
    resp = client.post("/movimentacoes", json=movement)
    assert resp.status_code == 400
    assert resp.json()["error"]["code"] == "validation_error"


def test_create_product_zero_stock_returns_friendly_message(client):
    resp = client.post("/produtos", json={"nome": "Produto Zero", "qtd_canoas": 0, "qtd_pf": 0})
    assert resp.status_code == 400
    body = resp.json()["error"]
    assert body["code"] == "validation_error"
    assert "Estoque inicial nao pode ser 0" in body["message"]


def test_analytics_saida_net_of_linked_devolucao(client):
    product_id = _create_product(client, "Produto Net", qtd_canoas=2, qtd_pf=0)
    date_saida = "2026-02-10T10:00:00"
    date_devolucao = "2026-02-10T12:00:00"

    saida = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 1,
            "origem": "CANOAS",
            "data": date_saida,
        },
    )
    assert saida.status_code == 201
    saida_id = saida.json()["data"]["id"]

    devolucao = client.post(
        "/movimentacoes",
        json={
            "tipo": "ENTRADA",
            "produto_id": product_id,
            "quantidade": 1,
            "destino": "CANOAS",
            "natureza": "DEVOLUCAO",
            "movimento_ref_id": saida_id,
            "data": date_devolucao,
        },
    )
    assert devolucao.status_code == 201

    top = client.get("/analytics/movements/top-saidas?date_from=2026-02-10&date_to=2026-02-10&scope=AMBOS")
    assert top.status_code == 200
    top_data = top.json()["data"]
    assert all(item["produto_id"] != product_id for item in top_data)

    ts = client.get("/analytics/movements/timeseries?date_from=2026-02-10&date_to=2026-02-10&scope=AMBOS&bucket=day")
    assert ts.status_code == 200
    ts_data = ts.json()["data"]
    assert sum(item["total_saida"] for item in ts_data) == 0

    flow = client.get("/analytics/movements/flow?date_from=2026-02-10&date_to=2026-02-10&scope=AMBOS&bucket=day")
    assert flow.status_code == 200
    flow_data = flow.json()["data"]
    assert len(flow_data) >= 1
    assert sum(item["saidas"] for item in flow_data) == 0


def test_list_movements_pagination(client):
    product_id = _create_product(client, "Produto List", qtd_canoas=1, qtd_pf=0)
    movement = {
        "tipo": "ENTRADA",
        "produto_id": product_id,
        "quantidade": 1,
        "destino": "PF",
    }
    client.post("/movimentacoes", json=movement)

    resp = client.get("/movimentacoes?page=1&page_size=1")
    assert resp.status_code == 200
    body = resp.json()
    assert body["data"][0]["produto_nome"] == "PRODUTO LIST"
    assert "meta" in body
    assert set(body["meta"].keys()) == {
        "page",
        "page_size",
        "total_items",
        "total_pages",
        "has_next",
    }


def test_backup_create(client):
    resp = client.post("/backup/criar")
    assert resp.status_code == 200
    body = resp.json()["data"]
    assert body["path"]
    assert body["size"] >= 0


def test_backup_list_and_validate(client):
    created = client.post("/backup/criar")
    assert created.status_code == 200
    created_path = created.json()["data"]["path"]
    created_name = os.path.basename(created_path)

    listed = client.get("/backup/listar")
    assert listed.status_code == 200
    items = listed.json()["data"]
    assert any(item["name"] == created_name for item in items)

    current_validation = client.get("/backup/validar")
    assert current_validation.status_code == 200
    assert bool(current_validation.json()["data"]["ok"]) is True

    selected_validation = client.get(f"/backup/validar?backup_name={created_name}")
    assert selected_validation.status_code == 200
    assert bool(selected_validation.json()["data"]["ok"]) is True


def test_backup_auto_config_and_restore_test(client):
    cfg = client.get("/backup/auto-config")
    assert cfg.status_code == 200
    assert "enabled" in cfg.json()["data"]
    assert cfg.json()["data"]["schedule_mode"] in {"DAILY", "WEEKLY"}
    assert 0 <= int(cfg.json()["data"]["weekday"]) <= 6

    updated = client.put(
        "/backup/auto-config",
        json={
            "enabled": True,
            "hour": 18,
            "minute": 0,
            "retention_days": 15,
            "schedule_mode": "WEEKLY",
            "weekday": 0,
        },
    )
    assert updated.status_code == 200
    assert bool(updated.json()["data"]["enabled"]) is True
    assert updated.json()["data"]["schedule_mode"] == "WEEKLY"
    assert updated.json()["data"]["weekday"] == 0

    created = client.post("/backup/criar")
    assert created.status_code == 200
    backup_name = os.path.basename(created.json()["data"]["path"])

    tested = client.post("/backup/testar-restauracao", json={"backup_name": backup_name})
    assert tested.status_code == 200
    assert bool(tested.json()["data"]["ok"]) is True


def test_backup_weekly_schedule_runs_only_on_configured_weekday(client):
    updated = client.put(
        "/backup/auto-config",
        json={
            "enabled": True,
            "hour": 18,
            "minute": 0,
            "retention_days": 7,
            "schedule_mode": "WEEKLY",
            "weekday": 0,
        },
    )
    assert updated.status_code == 200

    from app.services.backup_service import BackupService

    service = BackupService()

    tuesday = datetime(2026, 2, 17, 18, 5, 0)
    result_tuesday = service.run_due_scheduled_backup(tuesday)
    assert result_tuesday["executed"] is False
    assert result_tuesday["reason"] == "wrong_weekday"

    monday_before = datetime(2026, 2, 16, 17, 59, 0)
    result_before = service.run_due_scheduled_backup(monday_before)
    assert result_before["executed"] is False
    assert result_before["reason"] == "before_schedule"

    monday_due = datetime(2026, 2, 16, 18, 0, 0)
    result_due = service.run_due_scheduled_backup(monday_due)
    assert result_due["executed"] is True

    monday_repeat = datetime(2026, 2, 16, 18, 10, 0)
    result_repeat = service.run_due_scheduled_backup(monday_repeat)
    assert result_repeat["executed"] is False
    assert result_repeat["reason"] == "already_ran_today"


def test_backup_restore_roundtrip(client):
    product_id = _create_product(client, "Produto Restore", qtd_canoas=1, qtd_pf=0)

    snap = client.post("/backup/criar")
    assert snap.status_code == 200
    backup_name = os.path.basename(snap.json()["data"]["path"])

    movement = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 1,
            "origem": "CANOAS",
        },
    )
    assert movement.status_code == 201

    changed = client.get(f"/produtos/{product_id}")
    assert changed.status_code == 200
    assert changed.json()["data"]["qtd_canoas"] == 0

    restored = client.post("/backup/restaurar", json={"backup_name": backup_name})
    assert restored.status_code == 200

    after = client.get(f"/produtos/{product_id}")
    assert after.status_code == 200
    assert after.json()["data"]["qtd_canoas"] == 1


def test_backup_pre_update_restore(client):
    product_id = _create_product(client, "Produto PreUpdate", qtd_canoas=2, qtd_pf=0)

    pre = client.post("/backup/pre-update")
    assert pre.status_code == 200

    movement = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 1,
            "origem": "CANOAS",
        },
    )
    assert movement.status_code == 201

    restored = client.post("/backup/restaurar-pre-update")
    assert restored.status_code == 200

    after = client.get(f"/produtos/{product_id}")
    assert after.status_code == 200
    assert after.json()["data"]["qtd_canoas"] == 2


def test_backup_diagnostics_download(client):
    resp = client.get("/backup/diagnostico")
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith("application/zip")
    assert resp.content

    with zipfile.ZipFile(io.BytesIO(resp.content), "r") as zf:
        names = set(zf.namelist())
        assert "summary.json" in names
        assert "logs/backend.log.tail.txt" in names
        assert "logs/tauri.log.tail.txt" in names
        summary = json.loads(zf.read("summary.json").decode("utf-8"))
        assert "database" in summary
        assert "backups" in summary


def test_official_base_config_status_and_publish_apply(client, tmp_path):
    product_a = _create_product(client, "Produto Oficial A", qtd_canoas=2, qtd_pf=0)
    share_dir = tmp_path / "BaseOficialChronos"

    initial_status = client.get("/backup/base-oficial/status")
    assert initial_status.status_code == 200
    initial_data = initial_status.json()["data"]
    assert initial_data["role"] == "consumer"
    assert initial_data["latest_available"] is False

    config = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "publisher",
            "official_base_dir": str(share_dir),
            "machine_label": "PC-TESTE",
            "publisher_name": "Usuario Teste",
        },
    )
    assert config.status_code == 200
    config_data = config.json()["data"]
    assert config_data["role"] == "publisher"
    assert config_data["official_base_dir"] == str(share_dir)

    published = client.post(
        "/backup/base-oficial/publicar",
        json={"notes": "Base oficial apos conferencia"},
    )
    assert published.status_code == 200
    published_data = published.json()["data"]
    assert os.path.exists(published_data["zip_path"])
    assert os.path.exists(published_data["manifest_path"])
    assert os.path.exists(published_data["history_zip_path"])
    assert os.path.exists(published_data["history_manifest_path"])

    tested = client.post("/backup/base-oficial/testar-pasta")
    assert tested.status_code == 200
    tested_data = tested.json()["data"]
    assert tested_data["directory_exists"] is True
    assert tested_data["directory_accessible"] is True
    assert tested_data["read_ok"] is True
    assert tested_data["write_ok"] is True
    assert tested_data["latest_manifest_found"] is True

    history = client.get("/backup/base-oficial/historico?limit=5")
    assert history.status_code == 200
    history_data = history.json()["data"]
    assert len(history_data) >= 1
    assert history_data[0]["manifest"]["products_count"] == 1
    assert history_data[0]["manifest"]["movements_count"] >= 1

    product_b = _create_product(client, "Produto Oficial B", qtd_canoas=1, qtd_pf=0)
    listed_before_apply = client.get("/produtos")
    ids_before = {item["id"] for item in listed_before_apply.json()["data"]}
    assert product_a in ids_before
    assert product_b in ids_before

    applied = client.post("/backup/base-oficial/aplicar")
    assert applied.status_code == 200
    applied_data = applied.json()["data"]
    assert applied_data["restart_required"] is True
    assert os.path.exists(applied_data["pre_restore_backup"])

    listed_after_apply = client.get("/produtos")
    ids_after = {item["id"] for item in listed_after_apply.json()["data"]}
    assert product_a in ids_after
    assert product_b not in ids_after

    final_status = client.get("/backup/base-oficial/status")
    assert final_status.status_code == 200
    final_data = final_status.json()["data"]
    assert final_data["latest_available"] is True
    assert final_data["app_compatible_with_latest"] is True
    assert final_data["latest_manifest"]["publisher_machine"] == "PC-TESTE"
    assert final_data["latest_manifest"]["products_count"] == 1
    assert final_data["latest_manifest"]["movements_count"] >= 1


def test_official_base_apply_requires_compatible_app_version(client, tmp_path):
    _create_product(client, "Produto Compat", qtd_canoas=1, qtd_pf=0)
    share_dir = tmp_path / "BaseOficialChronos"

    configured = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "publisher",
            "official_base_dir": str(share_dir),
            "machine_label": "PC-TESTE",
        },
    )
    assert configured.status_code == 200

    published = client.post("/backup/base-oficial/publicar", json={})
    assert published.status_code == 200

    manifest_path = share_dir / "latest" / "base_oficial.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["min_app_version"] = "9.9.9"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    applied = client.post("/backup/base-oficial/aplicar")
    assert applied.status_code == 400
    assert "Atualize o aplicativo" in applied.json()["error"]["message"]


def test_official_base_can_delete_latest_and_history_publications(client, tmp_path):
    _create_product(client, "Produto Delete Base", qtd_canoas=1, qtd_pf=0)
    share_dir = tmp_path / "BaseOficialChronos"

    configured = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "publisher",
            "official_base_dir": str(share_dir),
            "machine_label": "PC-TESTE",
            "publisher_name": "Usuario Teste",
        },
    )
    assert configured.status_code == 200

    published = client.post("/backup/base-oficial/publicar", json={"notes": "Publicacao para excluir"})
    assert published.status_code == 200
    published_data = published.json()["data"]
    assert os.path.exists(published_data["manifest_path"])
    assert os.path.exists(published_data["zip_path"])

    history = client.get("/backup/base-oficial/historico?limit=5")
    assert history.status_code == 200
    history_data = history.json()["data"]
    assert len(history_data) == 1
    history_manifest_path = history_data[0]["manifest_path"]
    history_zip_path = history_data[0]["zip_path"]
    assert os.path.exists(history_manifest_path)
    assert os.path.exists(history_zip_path)

    deleted_history = client.request(
        "DELETE",
        "/backup/base-oficial/publicacoes",
        json={"manifest_path": history_manifest_path},
    )
    assert deleted_history.status_code == 200
    deleted_history_data = deleted_history.json()["data"]
    assert deleted_history_data["deleted_latest"] is False
    assert not os.path.exists(history_manifest_path)
    assert not os.path.exists(history_zip_path)

    status_before_latest_delete = client.get("/backup/base-oficial/status")
    assert status_before_latest_delete.status_code == 200
    assert status_before_latest_delete.json()["data"]["latest_available"] is True

    deleted_latest = client.request(
        "DELETE",
        "/backup/base-oficial/publicacoes",
        json={"delete_latest": True},
    )
    assert deleted_latest.status_code == 200
    deleted_latest_data = deleted_latest.json()["data"]
    assert deleted_latest_data["deleted_latest"] is True
    assert not os.path.exists(published_data["manifest_path"])
    assert not os.path.exists(published_data["zip_path"])

    final_status = client.get("/backup/base-oficial/status")
    assert final_status.status_code == 200
    assert final_status.json()["data"]["latest_available"] is False


def test_official_base_local_server_publish_and_apply(client):
    product_a = _create_product(client, "Produto Server A", qtd_canoas=2, qtd_pf=0)
    port = _find_free_port()
    server_url = f"http://127.0.0.1:{port}"

    configured = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "publisher",
            "machine_label": "PC-SERVIDOR",
            "publisher_name": "Usuario Teste",
            "server_port": port,
            "remote_server_url": server_url,
        },
    )
    assert configured.status_code == 200

    started = client.post("/backup/base-oficial-servidor/iniciar")
    assert started.status_code == 200
    assert started.json()["data"]["running"] is True

    try:
        published = client.post(
            "/backup/base-oficial-servidor/publicar",
            json={"notes": "Base oficial via servidor local"},
        )
        assert published.status_code == 200
        published_data = published.json()["data"]
        assert os.path.exists(published_data["zip_path"])
        assert os.path.exists(published_data["manifest_path"])

        remote = client.get(f"/backup/base-oficial-servidor/remoto?server_url={server_url}")
        assert remote.status_code == 200
        remote_data = remote.json()["data"]
        assert remote_data["official_available"] is True
        assert remote_data["machine_label"] == "PC-SERVIDOR"
        assert remote_data["official_manifest"]["products_count"] == 1

        product_b = _create_product(client, "Produto Server B", qtd_canoas=1, qtd_pf=0)
        listed_before_apply = client.get("/produtos")
        ids_before = {item["id"] for item in listed_before_apply.json()["data"]}
        assert product_a in ids_before
        assert product_b in ids_before

        applied = client.post(f"/backup/base-oficial-servidor/aplicar?server_url={server_url}")
        assert applied.status_code == 200
        applied_data = applied.json()["data"]
        assert applied_data["restart_required"] is True
        assert applied_data["publisher_machine"] == "PC-SERVIDOR"

        listed_after_apply = client.get("/produtos")
        ids_after = {item["id"] for item in listed_after_apply.json()["data"]}
        assert product_a in ids_after
        assert product_b not in ids_after
    finally:
        stopped = client.post("/backup/base-oficial-servidor/parar")
        assert stopped.status_code == 200
        assert stopped.json()["data"]["running"] is False


def test_compare_local_server_snapshot_flow(client):
    product_id = _create_product(client, "Produto Compare A", qtd_canoas=3, qtd_pf=0)
    port = _find_free_port()
    server_url = f"http://127.0.0.1:{port}"

    configured = client.put(
        "/backup/base-oficial/config",
        json={
            "role": "consumer",
            "machine_label": "PC-COMPARE",
            "server_port": port,
            "remote_server_url": server_url,
        },
    )
    assert configured.status_code == 200

    started = client.post("/backup/base-oficial-servidor/iniciar")
    assert started.status_code == 200

    try:
        published = client.post("/sistema/comparativo-servidor/publicar")
        assert published.status_code == 200

        remote = client.get(f"/sistema/comparativo-servidor/remoto?server_url={server_url}")
        assert remote.status_code == 200
        remote_data = remote.json()["data"]
        assert remote_data["compare_available"] is True
        assert remote_data["compare_manifest"]["total_items"] == 1

        moved = client.post(
            "/movimentacoes",
            json={"tipo": "ENTRADA", "produto_id": product_id, "quantidade": 2, "destino": "CANOAS"},
        )
        assert moved.status_code == 201

        compared = client.post("/sistema/comparativo-servidor/comparar", json={"server_url": server_url})
        assert compared.status_code == 200
        compare_data = compared.json()["data"]
        assert compare_data["summary"]["divergent_items"] >= 1
        assert compare_data["right"]["label"] == "Servidor remoto - PC-COMPARE"
    finally:
        stopped = client.post("/backup/base-oficial-servidor/parar")
        assert stopped.status_code == 200


def test_export_products(client):
    _create_product(client, "Produto Export", qtd_canoas=1, qtd_pf=0)
    resp = client.post("/export/produtos")
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.content


def test_export_stock_overview_excel(client):
    _create_product(client, "Produto Resumo A", qtd_canoas=2, qtd_pf=1)
    _create_product(client, "Produto Resumo B", qtd_canoas=0, qtd_pf=4)

    resp = client.post("/export/estoque-resumo")
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(io.BytesIO(resp.content))
    assert "Resumo" in workbook.sheetnames
    assert "Estoque" in workbook.sheetnames

    summary = workbook["Resumo"]
    stock = workbook["Estoque"]

    assert summary["A1"].value == "Chronos Inventory - Resumo de Estoque"
    assert summary["A4"].value == "Total de produtos ativos"
    assert summary["B4"].value >= 2
    assert summary["A8"].value == "Total global de pecas"
    assert summary["B8"].value >= 7

    assert stock["A1"].value == "ID"
    assert stock["B1"].value == "Produto"
    assert stock["C1"].value == "Estoque Canoas"
    assert stock["D1"].value == "Estoque PF"
    assert stock["E1"].value == "Total"
    assert stock["F1"].value == "Onde tem"
    assert stock.max_row >= 3


def test_report_selected_stock_pdf(client):
    product_a = _create_product(client, "Produto Selecionado A", qtd_canoas=2, qtd_pf=1)
    product_b = _create_product(client, "Produto Selecionado B", qtd_canoas=0, qtd_pf=4)
    _create_product(client, "Produto Fora Do Relatorio", qtd_canoas=7, qtd_pf=0)

    resp = client.post(
        "/relatorios/estoque-selecionado.pdf",
        json={"product_ids": [product_b, product_a]},
    )
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith("application/pdf")
    assert "Relatorio_Estoque_Selecionado.pdf" in (resp.headers.get("content-disposition") or "")
    assert resp.content


def test_report_real_sales_pdf(client):
    product_id = _create_product(client, "Produto Venda Relatorio", qtd_canoas=5, qtd_pf=0)
    movement = client.post(
        "/movimentacoes",
        json={
            "tipo": "SAIDA",
            "produto_id": product_id,
            "quantidade": 2,
            "origem": "CANOAS",
            "natureza": "OPERACAO_NORMAL",
            "documento": "NF 9988",
            "data": "2026-04-06T10:00:00",
        },
    )
    assert movement.status_code == 201

    resp = client.get("/relatorios/vendas-reais.pdf?date_from=2026-04-01&date_to=2026-04-07&scope=AMBOS")
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith("application/pdf")
    assert resp.content


def test_report_inactive_stock_pdf(client):
    _create_product(client, "Produto Parado Relatorio", qtd_canoas=3, qtd_pf=0)
    resp = client.get("/relatorios/estoque-parado.pdf?days=30&date_to=2026-04-07&scope=AMBOS")
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith("application/pdf")
    assert resp.content


def test_import_excel(client, tmp_path):
    df = pd.DataFrame([
        {"ID": 1, "Produto": "Item A", "Canoas": 1, "PF": 2},
    ])
    file_path = tmp_path / "import.xlsx"
    df.to_excel(file_path, index=False)

    with open(file_path, "rb") as f:
        files = {"file": ("import.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = client.post("/import/excel", files=files)

    assert resp.status_code == 200
    body = resp.json()["data"]
    assert "imported" in body


def test_dashboard_summary(client):
    _create_product(client, "Produto A", qtd_canoas=2, qtd_pf=3)
    product_b = _create_product(client, "Produto B", qtd_canoas=1, qtd_pf=0)
    movement = {
        "tipo": "SAIDA",
        "produto_id": product_b,
        "quantidade": 1,
        "origem": "CANOAS",
    }
    client.post("/movimentacoes", json=movement)

    resp = client.get("/dashboard/resumo")
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total_canoas"] == 2
    assert data["total_pf"] == 3
    assert data["total_geral"] == 5
    assert data["itens_distintos"] >= 2
    assert data["zerados"] >= 1


def test_upload_product_image(client):
    from PIL import Image

    product_id = _create_product(client, "Produto Imagem", qtd_canoas=1, qtd_pf=0)

    buffer = io.BytesIO()
    image = Image.new("RGB", (10, 10), color=(255, 0, 0))
    image.save(buffer, format="PNG")
    buffer.seek(0)

    files = {"file": ("produto.png", buffer, "image/png")}
    resp = client.post(f"/produtos/{product_id}/imagem", files=files)
    assert resp.status_code == 200
    body = resp.json()["data"]
    assert body["size_bytes"] > 0


def test_analytics_top_saidas_filters_and_order(client):
    product_a = _create_product(client, "Produto A", qtd_canoas=10, qtd_pf=0)
    product_b = _create_product(client, "Produto B", qtd_canoas=0, qtd_pf=5)
    product_c = _create_product(client, "Produto C", qtd_canoas=2, qtd_pf=0)
    product_d = _create_product(client, "Produto D", qtd_canoas=4, qtd_pf=0)

    base_date = "2026-02-10T10:00:00"

    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_a,
        "quantidade": 5,
        "origem": "CANOAS",
        "data": base_date,
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_b,
        "quantidade": 2,
        "origem": "PF",
        "data": base_date,
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_c,
        "quantidade": 2,
        "origem": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "MATRIZ",
        "data": base_date,
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_d,
        "quantidade": 1,
        "origem": "CANOAS",
        "natureza": "AJUSTE",
        "motivo_ajuste": "ERRO_OPERACIONAL",
        "observacao": "Ajuste teste",
        "data": base_date,
    })

    resp = client.get("/analytics/top-saidas?date_from=2026-02-10&date_to=2026-02-10&scope=AMBOS")
    assert resp.status_code == 200
    data = resp.json()["data"]
    ids = {item["produto_id"] for item in data}
    assert data[0]["produto_id"] == product_a
    assert data[0]["total_saida"] == 5
    assert product_b in ids
    assert product_c not in ids
    assert product_d not in ids

    resp_canoas = client.get("/analytics/top-saidas?date_from=2026-02-10&date_to=2026-02-10&scope=CANOAS")
    assert resp_canoas.status_code == 200
    data_canoas = resp_canoas.json()["data"]
    assert all(item["produto_id"] == product_a for item in data_canoas)

    resp_pf = client.get("/analytics/top-saidas?date_from=2026-02-10&date_to=2026-02-10&scope=PF")
    assert resp_pf.status_code == 200
    data_pf = resp_pf.json()["data"]
    assert all(item["produto_id"] == product_b for item in data_pf)


def test_analytics_recent_stockouts_only_considers_zero_stock_with_real_sales(client):
    product_recent = _create_product(client, "Produto Zerado Recente", qtd_canoas=3, qtd_pf=0)
    product_old = _create_product(client, "Produto Zerado Antigo", qtd_canoas=2, qtd_pf=0)
    product_transfer = _create_product(client, "Produto Zerado Transferencia", qtd_canoas=2, qtd_pf=0)
    product_not_zero = _create_product(client, "Produto Ainda Com Saldo", qtd_canoas=5, qtd_pf=0)
    product_pf = _create_product(client, "Produto PF Zerado", qtd_canoas=0, qtd_pf=2)

    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_recent,
        "quantidade": 3,
        "origem": "CANOAS",
        "data": "2026-02-10T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_old,
        "quantidade": 2,
        "origem": "CANOAS",
        "data": "2025-12-15T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_transfer,
        "quantidade": 2,
        "origem": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "MATRIZ",
        "data": "2026-02-10T11:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_not_zero,
        "quantidade": 2,
        "origem": "CANOAS",
        "data": "2026-02-10T12:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_pf,
        "quantidade": 2,
        "origem": "PF",
        "data": "2026-02-10T13:00:00",
    })

    resp_ambos = client.get("/analytics/products/recent-stockouts?days=30&date_to=2026-02-12&scope=AMBOS&limit=10")
    assert resp_ambos.status_code == 200
    data_ambos = resp_ambos.json()["data"]
    ids_ambos = {item["produto_id"] for item in data_ambos}
    assert product_recent in ids_ambos
    assert product_pf in ids_ambos
    assert product_old not in ids_ambos
    assert product_transfer not in ids_ambos
    assert product_not_zero not in ids_ambos

    resp_canoas = client.get("/analytics/products/recent-stockouts?days=30&date_to=2026-02-12&scope=CANOAS&limit=10")
    assert resp_canoas.status_code == 200
    data_canoas = resp_canoas.json()["data"]
    ids_canoas = {item["produto_id"] for item in data_canoas}
    assert product_recent in ids_canoas
    assert product_pf not in ids_canoas

    resp_pf = client.get("/analytics/products/recent-stockouts?days=30&date_to=2026-02-12&scope=PF&limit=10")
    assert resp_pf.status_code == 200
    data_pf = resp_pf.json()["data"]
    ids_pf = {item["produto_id"] for item in data_pf}
    assert product_pf in ids_pf
    assert product_recent not in ids_pf


def test_analytics_external_transfers_filters_type_scope_and_order(client):
    product_a = _create_product(client, "Produto Transferencia A", qtd_canoas=10, qtd_pf=0)
    product_b = _create_product(client, "Produto Transferencia B", qtd_canoas=0, qtd_pf=10)
    product_c = _create_product(client, "Produto Transferencia C", qtd_canoas=5, qtd_pf=0)

    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_a,
        "quantidade": 3,
        "origem": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "Matriz",
        "data": "2026-02-10T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_a,
        "quantidade": 2,
        "origem": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "Matriz",
        "data": "2026-02-11T09:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_b,
        "quantidade": 4,
        "origem": "PF",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "Canoas",
        "data": "2026-02-11T11:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "ENTRADA",
        "produto_id": product_b,
        "quantidade": 6,
        "destino": "PF",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "Matriz",
        "data": "2026-02-11T12:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "ENTRADA",
        "produto_id": product_c,
        "quantidade": 1,
        "destino": "CANOAS",
        "natureza": "TRANSFERENCIA_EXTERNA",
        "local_externo": "Filial",
        "data": "2026-02-12T08:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_c,
        "quantidade": 1,
        "origem": "CANOAS",
        "natureza": "OPERACAO_NORMAL",
        "data": "2026-02-12T10:00:00",
    })

    saidas = client.get(
        "/analytics/movements/external-transfers?date_from=2026-02-10&date_to=2026-02-12&scope=AMBOS&tipo=SAIDA&limit=10"
    )
    assert saidas.status_code == 200
    saidas_data = saidas.json()["data"]
    assert [item["produto_id"] for item in saidas_data[:2]] == [product_a, product_b]
    assert saidas_data[0]["total_quantidade"] == 5
    assert saidas_data[0]["total_movimentacoes"] == 2

    saidas_canoas = client.get(
        "/analytics/movements/external-transfers?date_from=2026-02-10&date_to=2026-02-12&scope=CANOAS&tipo=SAIDA&limit=10"
    )
    assert saidas_canoas.status_code == 200
    assert [item["produto_id"] for item in saidas_canoas.json()["data"]] == [product_a]

    entradas = client.get(
        "/analytics/movements/external-transfers?date_from=2026-02-10&date_to=2026-02-12&scope=AMBOS&tipo=ENTRADA&limit=10"
    )
    assert entradas.status_code == 200
    entradas_data = entradas.json()["data"]
    assert [item["produto_id"] for item in entradas_data[:2]] == [product_b, product_c]
    assert entradas_data[0]["total_quantidade"] == 6

    entradas_pf = client.get(
        "/analytics/movements/external-transfers?date_from=2026-02-10&date_to=2026-02-12&scope=PF&tipo=ENTRADA&limit=10"
    )
    assert entradas_pf.status_code == 200
    assert [item["produto_id"] for item in entradas_pf.json()["data"]] == [product_b]


def test_analytics_entradas_saidas_and_distribuicao(client):
    product_id = _create_product(client, "Produto Mov", qtd_canoas=1, qtd_pf=0)

    client.post("/movimentacoes", json={
        "tipo": "ENTRADA",
        "produto_id": product_id,
        "quantidade": 2,
        "destino": "CANOAS",
        "data": "2026-02-05T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_id,
        "quantidade": 1,
        "origem": "CANOAS",
        "data": "2026-02-06T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "TRANSFERENCIA",
        "produto_id": product_id,
        "quantidade": 1,
        "origem": "CANOAS",
        "destino": "PF",
        "data": "2026-02-06T12:00:00",
    })

    resp = client.get("/analytics/entradas-saidas?date_from=2026-02-05&date_to=2026-02-06")
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert len(data) == 2
    assert data[0]["entradas"] == 2
    assert data[1]["saidas"] == 1

    distrib = client.get("/analytics/estoque-distribuicao")
    assert distrib.status_code == 200
    distrib_data = distrib.json()["data"]
    assert distrib_data["total"] >= 0


def test_create_product_generates_initial_entry_in_flow(client):
    created = client.post(
        "/produtos",
        json={"nome": "Produto Fluxo Inicial", "qtd_canoas": 2, "qtd_pf": 3},
    )
    assert created.status_code == 201

    today = datetime.now().strftime("%Y-%m-%d")

    flow_ambos = client.get(
        f"/analytics/movements/flow?date_from={today}&date_to={today}&scope=AMBOS&bucket=day"
    )
    assert flow_ambos.status_code == 200
    flow_ambos_data = flow_ambos.json()["data"]
    assert len(flow_ambos_data) == 1
    assert flow_ambos_data[0]["entradas"] == 5
    assert flow_ambos_data[0]["saidas"] == 0

    flow_canoas = client.get(
        f"/analytics/movements/flow?date_from={today}&date_to={today}&scope=CANOAS&bucket=day"
    )
    assert flow_canoas.status_code == 200
    assert flow_canoas.json()["data"][0]["entradas"] == 2

    flow_pf = client.get(
        f"/analytics/movements/flow?date_from={today}&date_to={today}&scope=PF&bucket=day"
    )
    assert flow_pf.status_code == 200
    assert flow_pf.json()["data"][0]["entradas"] == 3


def test_analytics_top_sem_mov(client):
    product_id = _create_product(client, "Produto SemMov", qtd_canoas=1, qtd_pf=0)

    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_id,
        "quantidade": 1,
        "origem": "CANOAS",
        "data": "2025-12-01T10:00:00",
    })

    resp = client.get("/analytics/top-sem-mov?days=30&date_to=2026-02-10")
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert any(item["produto_id"] == product_id for item in data)


def test_product_images_multiple_flow(client):
    from PIL import Image

    product_id = _create_product(client, "Produto Multi Imagem", qtd_canoas=1, qtd_pf=0)

    def _img_bytes(color: tuple[int, int, int]) -> io.BytesIO:
        buffer = io.BytesIO()
        image = Image.new("RGB", (12, 12), color=color)
        image.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer

    files = [
        ("files", ("img1.png", _img_bytes((255, 0, 0)), "image/png")),
        ("files", ("img2.png", _img_bytes((0, 255, 0)), "image/png")),
    ]
    upload = client.post(f"/produtos/{product_id}/imagens", files=files)
    assert upload.status_code == 201
    assert upload.json()["data"]["total"] == 2

    listed = client.get(f"/produtos/{product_id}/imagens")
    assert listed.status_code == 200
    items = listed.json()["data"]["items"]
    assert len(items) == 2
    primary = [item for item in items if item["is_primary"]]
    assert len(primary) == 1

    second_id = [item["id"] for item in items if not item["is_primary"]][0]
    set_primary = client.patch(f"/produtos/{product_id}/imagens/{second_id}/principal")
    assert set_primary.status_code == 200

    listed_again = client.get(f"/produtos/{product_id}/imagens")
    assert listed_again.status_code == 200
    items_again = listed_again.json()["data"]["items"]
    assert any(item["id"] == second_id and item["is_primary"] for item in items_again)

    to_delete = items_again[0]["id"]
    deleted = client.delete(f"/produtos/{product_id}/imagens/{to_delete}")
    assert deleted.status_code == 200

    final_list = client.get(f"/produtos/{product_id}/imagens")
    assert final_list.status_code == 200
    assert final_list.json()["data"]["total"] == 1


def test_analytics_v2_endpoints(client):
    product_a = _create_product(client, "Produto V2 A", qtd_canoas=10, qtd_pf=0)
    product_b = _create_product(client, "Produto V2 B", qtd_canoas=5, qtd_pf=0)

    client.post("/movimentacoes", json={
        "tipo": "SAIDA",
        "produto_id": product_a,
        "quantidade": 3,
        "origem": "CANOAS",
        "data": "2026-02-10T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "ENTRADA",
        "produto_id": product_b,
        "quantidade": 4,
        "destino": "PF",
        "data": "2026-02-11T10:00:00",
    })
    client.post("/movimentacoes", json={
        "tipo": "TRANSFERENCIA",
        "produto_id": product_b,
        "quantidade": 2,
        "origem": "CANOAS",
        "destino": "PF",
        "data": "2026-02-11T12:00:00",
    })

    summary = client.get("/analytics/stock/summary")
    assert summary.status_code == 200
    summary_data = summary.json()["data"]
    assert "total_geral" in summary_data
    assert "zerados" in summary_data

    distribution = client.get("/analytics/stock/distribution")
    assert distribution.status_code == 200
    dist_data = distribution.json()["data"]
    assert "items" in dist_data
    assert len(dist_data["items"]) == 2

    top = client.get("/analytics/movements/top-saidas?date_from=2026-02-10&date_to=2026-02-12&scope=AMBOS")
    assert top.status_code == 200
    top_data = top.json()["data"]
    assert len(top_data) >= 1
    assert top_data[0]["produto_id"] == product_a

    ts = client.get("/analytics/movements/timeseries?date_from=2026-02-10&date_to=2026-02-12&scope=AMBOS&bucket=day")
    assert ts.status_code == 200
    assert isinstance(ts.json()["data"], list)

    flow = client.get("/analytics/movements/flow?date_from=2026-02-10&date_to=2026-02-12&scope=PF&bucket=day")
    assert flow.status_code == 200
    assert isinstance(flow.json()["data"], list)

    evolution = client.get("/analytics/stock/evolution?date_from=2026-02-10&date_to=2026-02-12&bucket=day")
    assert evolution.status_code == 200
    assert isinstance(evolution.json()["data"], list)

    inactive = client.get("/analytics/products/inactive?days=30&date_to=2026-02-12&limit=5")
    assert inactive.status_code == 200
    assert isinstance(inactive.json()["data"], list)


def test_inventory_session_flow(client):
    product_id = _create_product(client, "Produto Inventario", qtd_canoas=5, qtd_pf=0)
    _ = _create_product(client, "Produto Inventario PF", qtd_canoas=0, qtd_pf=2)

    created = client.post(
        "/inventario/sessoes",
        json={"nome": "Inventario Mensal", "local": "CANOAS", "observacao": "Conferencia geral"},
    )
    assert created.status_code == 201
    session = created.json()["data"]
    session_id = session["id"]

    listed = client.get("/inventario/sessoes")
    assert listed.status_code == 200
    assert any(item["id"] == session_id for item in listed.json()["data"])

    items = client.get(f"/inventario/sessoes/{session_id}/itens?only_divergent=false&page=1&page_size=100")
    assert items.status_code == 200
    rows = items.json()["data"]
    target = next((row for row in rows if row["produto_id"] == product_id), None)
    assert target is not None

    update = client.put(
        f"/inventario/sessoes/{session_id}/itens",
        json={
            "items": [
                {
                    "produto_id": product_id,
                    "qtd_fisico": 3,
                    "motivo_ajuste": "CORRECAO_INVENTARIO",
                    "observacao": "Contagem fisica final",
                }
            ]
        },
    )
    assert update.status_code == 200

    applied = client.post(f"/inventario/sessoes/{session_id}/aplicar")
    assert applied.status_code == 200
    assert applied.json()["data"]["applied_items"] >= 1

    product = client.get(f"/produtos/{product_id}")
    assert product.status_code == 200
    assert product.json()["data"]["qtd_canoas"] == 3


def test_inventory_session_summary_and_filters(client):
    missing_id = _create_product(client, "Produto Falta", qtd_canoas=5, qtd_pf=0)
    surplus_id = _create_product(client, "Produto Sobra", qtd_canoas=1, qtd_pf=0)
    matched_id = _create_product(client, "Produto OK", qtd_canoas=2, qtd_pf=0)

    created = client.post(
        "/inventario/sessoes",
        json={"nome": "Inventario Divergencias", "local": "CANOAS"},
    )
    assert created.status_code == 201
    session_id = created.json()["data"]["id"]

    update = client.put(
        f"/inventario/sessoes/{session_id}/itens",
        json={
            "items": [
                {
                    "produto_id": missing_id,
                    "qtd_fisico": 3,
                    "motivo_ajuste": "CORRECAO_INVENTARIO",
                    "observacao": "Faltou no fisico",
                },
                {
                    "produto_id": surplus_id,
                    "qtd_fisico": 4,
                    "motivo_ajuste": "CORRECAO_INVENTARIO",
                    "observacao": "Sobrou no fisico",
                },
                {
                    "produto_id": matched_id,
                    "qtd_fisico": 2,
                    "observacao": "Conferido",
                },
            ]
        },
    )
    assert update.status_code == 200

    summary = client.get(f"/inventario/sessoes/{session_id}/resumo")
    assert summary.status_code == 200
    summary_data = summary.json()["data"]
    assert summary_data["missing_items"] >= 1
    assert summary_data["surplus_items"] >= 1
    assert summary_data["matched_items"] >= 1
    assert summary_data["pending_items"] >= 2

    missing = client.get(f"/inventario/sessoes/{session_id}/itens?status_filter=MISSING&page=1&page_size=100")
    assert missing.status_code == 200
    missing_ids = {row["produto_id"] for row in missing.json()["data"]}
    assert missing_id in missing_ids
    assert surplus_id not in missing_ids

    surplus = client.get(f"/inventario/sessoes/{session_id}/itens?status_filter=SURPLUS&page=1&page_size=100")
    assert surplus.status_code == 200
    surplus_ids = {row["produto_id"] for row in surplus.json()["data"]}
    assert surplus_id in surplus_ids
    assert missing_id not in surplus_ids

    matched = client.get(f"/inventario/sessoes/{session_id}/itens?status_filter=MATCHED&page=1&page_size=100")
    assert matched.status_code == 200
    matched_ids = {row["produto_id"] for row in matched.json()["data"]}
    assert matched_id in matched_ids


def test_inventory_session_can_be_closed_and_deleted(client):
    product_id = _create_product(client, "Produto Inventario Fechar", qtd_canoas=2, qtd_pf=0)

    created = client.post(
        "/inventario/sessoes",
        json={"nome": "Inventario para fechar", "local": "CANOAS"},
    )
    assert created.status_code == 201
    session_id = created.json()["data"]["id"]

    closed = client.post(f"/inventario/sessoes/{session_id}/fechar")
    assert closed.status_code == 200
    assert closed.json()["data"]["status"] == "FECHADO"

    update = client.put(
        f"/inventario/sessoes/{session_id}/itens",
        json={
            "items": [
                {
                    "produto_id": product_id,
                    "qtd_fisico": 1,
                    "motivo_ajuste": "CORRECAO_INVENTARIO",
                    "observacao": "teste",
                }
            ]
        },
    )
    assert update.status_code == 400

    deleted = client.delete(f"/inventario/sessoes/{session_id}")
    assert deleted.status_code == 200
    assert deleted.json()["data"]["session_id"] == session_id

    missing = client.get(f"/inventario/sessoes/{session_id}")
    assert missing.status_code == 400


def test_inventory_applied_session_cannot_be_deleted(client):
    product_id = _create_product(client, "Produto Inventario Aplicado", qtd_canoas=5, qtd_pf=0)

    created = client.post(
        "/inventario/sessoes",
        json={"nome": "Inventario aplicado", "local": "CANOAS"},
    )
    assert created.status_code == 201
    session_id = created.json()["data"]["id"]

    updated = client.put(
        f"/inventario/sessoes/{session_id}/itens",
        json={
            "items": [
                {
                    "produto_id": product_id,
                    "qtd_fisico": 3,
                    "motivo_ajuste": "CORRECAO_INVENTARIO",
                    "observacao": "fechar depois",
                }
            ]
        },
    )
    assert updated.status_code == 200

    applied = client.post(f"/inventario/sessoes/{session_id}/aplicar")
    assert applied.status_code == 200

    deleted = client.delete(f"/inventario/sessoes/{session_id}")
    assert deleted.status_code == 400
    assert deleted.json()["error"]["code"] == "validation_error"


def test_inventory_session_does_not_include_inactive_products(client):
    active_id = _create_product(client, "Produto Inventario Ativo", qtd_canoas=4, qtd_pf=0)
    inactive_id = _create_product(client, "Produto Inventario Inativo", qtd_canoas=7, qtd_pf=0)

    toggle = client.put(
        "/produtos/status-lote",
        json={"ids": [inactive_id], "ativo": False, "motivo_inativacao": "Obsoleto"},
    )
    assert toggle.status_code == 200

    created = client.post(
        "/inventario/sessoes",
        json={"nome": "Inventario Ativos", "local": "CANOAS", "observacao": "Somente ativos"},
    )
    assert created.status_code == 201
    session_id = created.json()["data"]["id"]

    items = client.get(f"/inventario/sessoes/{session_id}/itens?only_divergent=false&page=1&page_size=200")
    assert items.status_code == 200
    ids = {row["produto_id"] for row in items.json()["data"]}
    assert active_id in ids
    assert inactive_id not in ids
